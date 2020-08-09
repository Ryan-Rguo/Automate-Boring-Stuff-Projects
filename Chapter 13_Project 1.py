# Chapter 13
# Practice Project 1

# Create X * Y multiplication table in Excel.

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def multiTable():
     print('To create a X * Y multiplication table in Excel.')
     print('please input X')
     X = int(input())
     print('please input Y')
     Y = int(input())
      
     wb = openpyxl.Workbook()
     boldFont = Font(bold = True)
     sheet = wb.active

     for i in range(X):
          columnCell = sheet.cell(row = 1,column = i + 2)
          columnCell.value = i + 1
          columnCell.font = boldFont

     for v in range(Y):
          rowCell = sheet.cell(row = v + 2,column = 1)
          rowCell.value = v + 1
          rowCell.font = boldFont

     for rowNum in range(2,Y + 2):
          for columnNum in range(2,X + 2):
               columnLetter = get_column_letter(columnNum)
               sheet.cell(row = rowNum, column = columnNum).value = '=%s%s*%s%s' % ('A',str(rowNum),columnLetter,'1')

     wb.save('Multiplication Table.xlsx')
     print('Done')
          
          
