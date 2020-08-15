# Chapter 16
# Practice Project 1

# Excel to CSV Converter: To convert each sheet from Excel files to CSV.

import openpyxl, csv, os

os.makedirs('CSV Files',exist_ok = True)

def excelToCsv():
    for excelFile in os.listdir('.'):
        if not excelFile.endswith('.xlsx'):
            continue
        wb = openpyxl.load_workbook(excelFile)

        for sheetName in wb.sheetnames:
            sheet = wb[sheetName]
            csvFileName = excelFile.rstrip('.xlsx') + '_' + sheetName + '.csv'
            outputFile = open(os.path.join('CSV Files',csvFileName), 'w', newline = '')
            csvWriter = csv.writer(outputFile)

            for row in range(sheet.max_row):
                rowContent = []
                for column in range(sheet.max_column):
                    cell = sheet.cell(row = row + 1, column = column + 1).value
                    rowContent.append(cell)
                csvWriter.writerow(rowContent)

        outputFile.close()

    print('Done')

excelToCsv()