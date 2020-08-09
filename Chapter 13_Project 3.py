# Chapter 13
# Practice Project 3

# Spreadsheet Cell Inverter
# In this project, a new sheet is created and inverted content is copied there.

import openpyxl

def cellInverter():

     wb = openpyxl.load_workbook('project.xlsx')  
     wb.create_sheet(index=1,title='Sheet2')

     sheet1 = wb['Sheet']
     sheet2 = wb['Sheet2']

     for rowNum in range(1,sheet1.max_row + 1):
          for columnNum in range(1,sheet1.max_column + 1):
               sheet2.cell(row = columnNum, column = rowNum).value = sheet1.cell(row = rowNum, column = columnNum).value


     wb.save('project_inverted.xlsx')
     print('Done')
