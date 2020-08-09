# Chapter 13
# Practice Project 5

# Spreadsheet to Text Files

import openpyxl
from openpyxl.utils import get_column_letter


def excelToTxt():
     wb = openpyxl.load_workbook('Multiplication Table.xlsx')

     sheet = wb.active

     for columnNum in range(1,sheet.max_column + 1):
          file1 = open('excel_%s.txt' % get_column_letter(columnNum), 'w')
          for rowNum in range(1,sheet.max_row + 1):
               copy = sheet.cell(row = rowNum, column = columnNum).value
               if copy is not None:
                    file1.write(str(copy) + '\n') 
               else:
                    copy = '\n'
                    file1.write(copy)
          file1.close()

     print('Done')
