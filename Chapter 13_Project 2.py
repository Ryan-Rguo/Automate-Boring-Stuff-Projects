# Chapter 13
# Practice Project 2

# Blank Row Inserter
# Usage - Insert m rows in row number n.
# In this project, the name of .bat file is 'excelInserter.bat'
# Input 'excelInserter n m project.xlsx' in 'run'

import openpyxl, sys
from openpyxl.styles import Font

wb = openpyxl.load_workbook(sys.argv[3])
wb.create_sheet(index=1,title='Sheet2')

n = int(sys.argv[1])
m = int(sys.argv[2])

sheet1 = wb['Sheet1']
sheet2 = wb['Sheet2']

for rowNum in range(1,n):
     for columnNum in range(1,sheet1.max_column + 1):
          if rowNum == 1:
               sheet2.cell(row = rowNum, column = columnNum).font = Font(bold = True)
          sheet2.cell(row = rowNum, column = columnNum).value = sheet1.cell(row = rowNum, column = columnNum).value


for rowNum in range(n,sheet1.max_row + 1):
     for columnNum in range(1,sheet1.max_column + 1):
          sheet2.cell(row = rowNum + m, column = columnNum).value = sheet1.cell(row = rowNum, column = columnNum).value
wb.save('duesRecords_updated.xlsx')
print('Done')
