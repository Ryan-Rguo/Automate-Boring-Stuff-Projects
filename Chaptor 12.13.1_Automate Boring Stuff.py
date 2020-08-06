# 12.13.1
# Create X * Y multiplication table in Excel.

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

print('To create a X * Y multiplication table in Excel.')
print('please input X')
X = int(input())
print('please input Y')
Y = int(input())
      
wb = openpyxl.Workbook()
boldFont = Font(bold = True)
sheet = wb.active

for i in range(2,X + 2):
     columnCell = sheet.cell(row = 1,column = i)
     columnCell.value = i - 1
     columnCell.font = boldFont

for v in range(2,Y + 2):
     rowCell = sheet.cell(row = v,column = 1)
     rowCell.value = v - 1
     rowCell.font = boldFont

for rowNum in range(2,Y + 2):
     for columnNum in range(2,X + 2):
          columnLetter = get_column_letter(columnNum)
          sheet.cell(row = rowNum, column = columnNum).value = '=%s%s*%s%s' % ('A',str(rowNum),columnLetter,'1')

wb.save('Multiplication Table.xlsx')
print('Done')
          
          
